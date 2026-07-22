"""Dependency-light rewrite of catmodeling.py.

Same behavior and output as catmodeling.py, but without pandas/numpy (and without the
dead python_calamine import). All DataFrame-style operations are re-expressed with
plain Python (lists of dicts + an ordered column-name list) and the stdlib csv module.
openpyxl is kept -- it's the only library that can copy the source workbook's original
formatting verbatim and then append newly-styled sheets to it.
"""

import csv
import datetime
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

base_path = Path.cwd()
excel_files = list((base_path / "data").glob("*.xlsx"))
if len(excel_files) != 1:
    raise SystemExit(f"Expected exactly one .xlsx in data/, found: {excel_files}")
raw_data = excel_files[0]

RENAME_COLS = {
    'MCONAME': 'MCONAME',
    'Agent Name': 'Agent Name',
    'Named Insured': 'Named Insured',  # get ACCNTNAME from name insured with commas stripped for proper CSV parsing
    'QUOTEID': 'QUOTEID',  # ACCNTNUM gets value from this column
    'LOCNO': 'LOCNO',
    'BLDGNO': 'BLDGNO',    # get LOCNUM column from this column
    'STNAME': 'STNAME',    # get STREETNAME column from this column
    'CITY': 'CITY',        # duplicate CITY column? --this probably matters when we get only the yellow columns for the CSV output so that the CSV parses correctly when inputted into the model
    'STATE': 'STATE',      # get STATECODE column from this column
    'ZIP5': 'ZIP5',        # get POSTALCODE column from this column
    'COUNTY': 'COUNTY',    # duplicate COUNTY column; other columns: CNTRYSCHEME == ISO2A, CNTRYCODE == US
    'CONSTCL': 'CONSTCL',  # after this we have BLDGSCHEME == FIRE and BLDGCLASS which comes from this column
    'OCCPCL': 'OCCPCL',    # after this we have OCCSCHEME == ATC and OCCTYPE which comes from this column
    'NOSTORIES': 'NOSTORIES',  # get NUMSTORIES column from this column
    'YEARBUILT': 'YEARBUILT',  # duplicate YEARBUILT column
    'RMSLOB': 'RMSLOB',
    'WINDPCTDED': 'WINDPCTDED',
    'LOCBLKLIMIT': 'LOCBLKLIMIT',
    'LOCBLKDEDU': 'LOCBLKDEDU',
    'LOCBLDREPL': 'LOCBLDREPL',    # get CV1VAL column from this column
    'LOCBLDLIMT': 'LOCBLDLIMT',
    'LOCBLDDEDU': '*LOCBLDDEDU',
    'LOCCNTREPL': 'LOCCNTREPL',    # get CV2VAL column from this column
    'COCCNTLIMT': 'COCCNTLIMT',
    'LOCCNTDEDU': 'LOCCNTDEDU',
    'LOCBIRC': 'LOCBIRC',
    'LOCBILIMIT': 'LOCBILIMIT',
    'LOCBIDEDU': 'LOCBIDEDU',
    'LOCBLDPREMAOP': 'LOCBLDPREMAOP',
    'LOCBLDPREMWIND': 'LOCBLDPREMWIND',
    'LOCCNTPREMAOP': 'LOCCNTPREMAOP',
    'LOCCNTPREMWIND': 'LOCCNTPREMWIND',
    'LOCBIPREM': 'LOCBIPREM',
    'SQFEET': 'SQFEET',    # get FLOORAREA column from this column
    'RATINGTERR': 'RATINGTERR',
    'SPRINKLER': 'SPRINKLER',
    'PROTCLASS': 'PROTCLASS',
    'DESCRIPTION': 'DESCRIPTION',
    'Effective Date': 'Effective Date',    # get INCEPTDATE column from this column
    'Expiration Date': 'Expiration Date',  # get EXPIREDATE column from this column
    'CONSTQUA': '*CONSTQUALI',
    'ROOFSYS': '*ROOFSYS',
    'ROOFAGE': '*ROOFAGE',
    'ROOFGEO': '*ROOFGEOM',
    'ROOFANC': '*ROOFANCH',
    'CLADSYS': '*CLADSYS',
    'FOUNDSYS': '*FOUNDSYS',
    'ROOFEQUI': '*ROOFEQUIP',
    'CLADRATE': '*CLADRATE',
    'RESISTOPEN': '*RESISTOPEN'
}

SCENARIOS = [
    ('2pct', 'SCS', 3, 0.02),
    ('2pct', 'WS', 2, 0.02),
    ('3pct', 'WS', 2, 0.03),
    ('5pct', 'WS', 2, 0.05)
]


# ---------------------------------------------------------------------------
# Minimal helpers standing in for a DataFrame: an ordered `columns` name list
# plus `rows` as a list of dicts keyed by column name. `insert_col` mirrors
# pandas' df.insert(loc, name, value): splice the name into the column list at
# the given position and set the value on every row. Column position lookups
# use plain list.index(), mirroring df.columns.get_loc(name). Kept intentionally
# minimal -- just enough to replicate the exact sequence of operations below.
# ---------------------------------------------------------------------------

def get_loc(columns, name):
    return columns.index(name)


def insert_col(columns, rows, index, name, value):
    """Mirrors df.insert(index, name, value). `value` may be a constant or a
    callable(row) -> value computed per-row."""
    columns.insert(index, name)
    if callable(value):
        for row in rows:
            row[name] = value(row)
    else:
        for row in rows:
            row[name] = value


def drop_cols(columns, rows, names):
    for name in names:
        columns.remove(name)
    for row in rows:
        for name in names:
            row.pop(name, None)


def project(rows, columns):
    """Select only `columns` from each row (mirrors df[[...]])."""
    return [{c: row.get(c) for c in columns} for row in rows]


def dedupe_first(rows, key_cols):
    """Keeps the first row seen per unique key, preserving order (mirrors
    drop_duplicates(subset=key_cols), which keeps the literal first row)."""
    seen = set()
    out = []
    for row in rows:
        key = tuple(row[c] for c in key_cols)
        if key not in seen:
            seen.add(key)
            out.append(row)
    return out


def groupby_first(rows, key_cols, columns):
    """Mirrors df.groupby(key_cols, as_index=False, sort=False).first():
    NOT a literal "keep the first row" -- pandas' .first() takes, independently
    for each column, the first non-null value seen across all rows sharing the
    group key. Two rows in the same group can each contribute different
    columns to the merged output row. (Verified against a live pandas run:
    groupby().first() interleaves non-null values column-by-column rather than
    picking one whole row.)"""
    merged_by_key = {}
    order = []
    for row in rows:
        key = tuple(row[c] for c in key_cols)
        if key not in merged_by_key:
            merged_by_key[key] = dict(row)
            order.append(key)
        else:
            merged = merged_by_key[key]
            for c in columns:
                if merged.get(c) is None and row.get(c) is not None:
                    merged[c] = row[c]
    return [merged_by_key[k] for k in order]


def num_to_str(value):
    """str(value), but a float with no fractional part is converted to int
    first. pandas' .astype(str) on a float column yields "2005.0"; openpyxl
    returns whole numbers as int, so plain str() would give "2005" already --
    this guards the reverse case (a float slipping in) so string-built columns
    like *YEARBUILT / *ACCNTNUM never pick up a stray ".0"."""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def classify_column(values):
    """Approximates the dtype pandas would infer for a column built the way
    this script builds one: read straight from Excel cells, then unioned via
    concatenation of several per-scenario tables. If every non-null value is
    an integer (or an integer-looking string) and there is no null anywhere,
    pandas keeps it int64. If there's a null mixed in with otherwise-clean
    numbers (from a source column, or from unioning a None-only scenario
    branch with an int-only one), the whole column is forced to float64 --
    even the integer-valued rows then render as "10000.0". Any value that
    isn't numeric-like at all leaves the column untouched ("other")."""
    saw_none = False
    saw_non_integer_number = False
    for v in values:
        if v is None:
            saw_none = True
            continue
        if isinstance(v, bool):
            return 'other'
        if isinstance(v, int):
            continue
        if isinstance(v, float):
            if not v.is_integer():
                saw_non_integer_number = True
            continue
        if isinstance(v, str):
            s = v.strip()
            if s == '':
                return 'other'
            try:
                int(s)
                continue
            except ValueError:
                pass
            try:
                float(s)
                saw_non_integer_number = True
                continue
            except ValueError:
                return 'other'
            continue
        return 'other'
    if saw_none or saw_non_integer_number:
        return 'float'
    return 'int'


def coerce_column_types(columns, rows):
    """Applies classify_column's verdict to every value in each column, in
    place. This is what reproduces pandas' whole-column dtype promotion (see
    classify_column) at the point each of the three final tables is complete,
    since that's the only time the *entire* column (across every source row
    and every concatenated scenario) is known."""
    for col in columns:
        values = [row.get(col) for row in rows]
        kind = classify_column(values)
        if kind == 'int':
            for row in rows:
                v = row.get(col)
                if v is not None:
                    row[col] = int(v)
        elif kind == 'float':
            for row in rows:
                v = row.get(col)
                if v is not None:
                    row[col] = float(v)


# ---------------------------------------------------------------------------
# 1-2. Read the workbook, drop the blank/Unnamed trailing column.
# ---------------------------------------------------------------------------

wb_raw = openpyxl.load_workbook(raw_data, read_only=True, data_only=True)
ws_raw = wb_raw.worksheets[0]
raw_iter = ws_raw.iter_rows(values_only=True)
raw_header = list(next(raw_iter))
raw_data_rows = [list(r) for r in raw_iter]
wb_raw.close()

blank_idx = next(i for i, h in enumerate(raw_header) if h is None or str(h).strip() == '')
del raw_header[blank_idx]
for r in raw_data_rows:
    del r[blank_idx]

# ---------------------------------------------------------------------------
# 3. Rename columns via RENAME_COLS (build modeling's columns/rows directly
#    under the renamed keys).
# ---------------------------------------------------------------------------

modeling_columns = [RENAME_COLS.get(c, c) for c in raw_header]
modeling_rows = [
    {RENAME_COLS.get(raw_header[i], raw_header[i]): v for i, v in enumerate(r)}
    for r in raw_data_rows
]

# ---------------------------------------------------------------------------
# 4. Add the * columns -- replicate every insert() from the reference script,
#    in the same order (each get_loc is recomputed after prior inserts).
# ---------------------------------------------------------------------------

insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'Named Insured') + 1,
           '*ACCNTNAME', lambda row: row['Named Insured'].replace(',', '') if isinstance(row['Named Insured'], str) else row['Named Insured'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'QUOTEID') + 1,
           '*ACCNTNUM', lambda row: row['QUOTEID'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'BLDGNO') + 1,
           '*LOCNUM', lambda row: row['BLDGNO'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'COUNTY') + 1,
           '*STREETNAME', lambda row: row['STNAME'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'COUNTY') + 2,
           '*CITY', lambda row: row['CITY'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'COUNTY') + 3,
           '*STATECODE', lambda row: row['STATE'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'COUNTY') + 4,
           '*POSTALCODE', lambda row: row['ZIP5'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'COUNTY') + 5,
           '*COUNTY', lambda row: row['COUNTY'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'COUNTY') + 6,
           '*CNTRYSCHEME', 'ISO2A')
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'COUNTY') + 7,
           '*CNTRYCODE', 'US')
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'CONSTCL') + 1,
           '*BLDGSCHEME', 'FIRE')
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'CONSTCL') + 2,
           '*BLDGCLASS', lambda row: row['CONSTCL'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'OCCPCL') + 1,
           '*OCCSCHEME', 'ATC')
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'OCCPCL') + 2,
           '*OCCTYPE', lambda row: row['OCCPCL'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'NOSTORIES') + 1,
           '*NUMSTORIES', lambda row: row['NOSTORIES'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'YEARBUILT') + 1,
           '*YEARBUILT', lambda row: '1/1/' + num_to_str(row['YEARBUILT']))
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'LOCBLDREPL'),
           '*CV1VAL', lambda row: row['LOCBLDREPL'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'LOCCNTREPL'),
           '*CV2VAL', lambda row: row['LOCCNTREPL'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'SQFEET') + 1,
           '*FLOORAREA', lambda row: row['SQFEET'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'Expiration Date') + 1,
           '*INCEPTDATE', lambda row: row['Effective Date'])
insert_col(modeling_columns, modeling_rows, get_loc(modeling_columns, 'Expiration Date') + 2,
           '*EXPIREDATE', lambda row: row['Expiration Date'])

# ---------------------------------------------------------------------------
# 5-6. acct_temp / location_temp.
# ---------------------------------------------------------------------------

acct_temp_columns = ['*ACCNTNUM', '*ACCNTNAME', '*INCEPTDATE', '*EXPIREDATE', '*LOCBLDDEDU']
acct_temp_rows = dedupe_first(project(modeling_rows, acct_temp_columns), ['*ACCNTNUM'])

location_temp_columns = [c for c in modeling_columns if c.startswith('*')]
location_temp_rows = project(modeling_rows, location_temp_columns)

# ---------------------------------------------------------------------------
# 7. Scenario loop.
# ---------------------------------------------------------------------------

acct_scenario_tables = []
loc_scenario_tables = []

for pct, pnum, ptype, prob in SCENARIOS:
    acct_columns = list(acct_temp_columns)
    acct_rows = [dict(row) for row in acct_temp_rows]
    for row in acct_rows:
        row['*ACCNTNUM'] = num_to_str(row['*ACCNTNUM']) + '_' + pct
    idx = get_loc(acct_columns, '*EXPIREDATE')
    insert_col(acct_columns, acct_rows, idx + 1, '*POLICYNUM', pnum)
    insert_col(acct_columns, acct_rows, idx + 2, '*POLICYTYPE', ptype)

    loc_columns = list(location_temp_columns)
    loc_rows = [dict(row) for row in location_temp_rows]
    for row in loc_rows:
        row['*ACCNTNUM'] = num_to_str(row['*ACCNTNUM']) + '_' + pct
    insert_col(loc_columns, loc_rows, get_loc(loc_columns, '*CV1VAL') + 1, '*WSSITEDED', prob)

    if pnum == 'SCS':
        for row in acct_rows:
            row['*BLANDEDAMT'] = row['*LOCBLDDEDU']
            row['*MINDEDAMT'] = None
        acct_columns.append('*BLANDEDAMT')
        acct_columns.append('*MINDEDAMT')

        idx2 = get_loc(loc_columns, '*CV2VAL')
        insert_col(loc_columns, loc_rows, idx2 + 1, '*TOCV1VAL', lambda row: row['*CV1VAL'])
        insert_col(loc_columns, loc_rows, idx2 + 2, '*TOCV2VAL', lambda row: row['*CV2VAL'])
        idx1 = get_loc(loc_columns, '*CV1VAL')
        insert_col(loc_columns, loc_rows, idx1 + 1, '*WSCV1VAL', None)
        insert_col(loc_columns, loc_rows, idx1 + 2, '*WSCV2VAL', None)
    else:
        for row in acct_rows:
            row['*BLANDEDAMT'] = None
            row['*MINDEDAMT'] = row['*LOCBLDDEDU']
        acct_columns.append('*BLANDEDAMT')
        acct_columns.append('*MINDEDAMT')

        idx2 = get_loc(loc_columns, '*CV2VAL')
        insert_col(loc_columns, loc_rows, idx2 + 1, '*TOCV1VAL', None)
        insert_col(loc_columns, loc_rows, idx2 + 2, '*TOCV2VAL', None)
        idx1 = get_loc(loc_columns, '*CV1VAL')
        insert_col(loc_columns, loc_rows, idx1 + 1, '*WSCV1VAL', lambda row: row['*CV1VAL'])
        insert_col(loc_columns, loc_rows, idx1 + 2, '*WSCV2VAL', lambda row: row['*CV2VAL'])

    drop_cols(loc_columns, loc_rows, ['*CV1VAL', '*CV2VAL'])

    acct_scenario_tables.append((acct_columns, acct_rows))
    loc_scenario_tables.append((loc_columns, loc_rows))

# ---------------------------------------------------------------------------
# 8. Concatenate the 4 account tables and the 4 location tables.
# ---------------------------------------------------------------------------

account_columns = list(acct_scenario_tables[0][0])
account_columns.remove('*LOCBLDDEDU')
account_rows = []
for cols, rows in acct_scenario_tables:
    account_rows.extend(project(rows, account_columns))

location_columns = list(loc_scenario_tables[0][0])
location_columns.remove('*LOCBLDDEDU')
location_rows = []
for cols, rows in loc_scenario_tables:
    location_rows.extend(project(rows, location_columns))

# Reproduce pandas' whole-column dtype promotion (see classify_column) on the
# concatenated-but-not-yet-deduped table. This has to happen *before* the
# groupby-first merge below: pandas' concat is what forces a column like
# *WSCV1VAL to float64 in the first place (the SCS scenario's block is
# all-null there while the WS scenarios' blocks are clean ints), and that
# float64 dtype is "sticky" -- it survives groupby().first() even for groups
# where every null in that column happens to get filled in during the merge.
# Coercing only the post-dedupe result would miss this, since by then no None
# may be left in the column to signal that promotion should have happened.
coerce_column_types(account_columns, account_rows)
coerce_column_types(location_columns, location_rows)

# ---------------------------------------------------------------------------
# 9. *ACCNTNAME/*ACCNTNUM column swap + group-first dedupe on (*ACCNTNUM, *LOCNUM).
# ---------------------------------------------------------------------------

name_idx = location_columns.index('*ACCNTNAME')
num_idx = location_columns.index('*ACCNTNUM')
location_columns[name_idx], location_columns[num_idx] = location_columns[num_idx], location_columns[name_idx]

location_rows = groupby_first(location_rows, ['*ACCNTNUM', '*LOCNUM'], location_columns)

coerce_column_types(modeling_columns, modeling_rows)

# ---------------------------------------------------------------------------
# 10. Excel output: copy the source workbook verbatim (preserves its original
#     formatting), then append the derived sheets with yellow-filled headers
#     for columns that were originally `*`-prefixed.
# ---------------------------------------------------------------------------

output_dir = base_path / "output"
output_dir.mkdir(exist_ok=True)
output_path = output_dir / "WindModeling_Output.xlsx"

shutil.copyfile(raw_data, output_path)

derived_sheets = [
    ("ModelingSOV", modeling_columns, modeling_rows),
    ("Account", account_columns, account_rows),
    ("Location", location_columns, location_rows),
]

wb_out = openpyxl.load_workbook(output_path)
yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
for sheet_name, cols, rows in derived_sheets:
    ws = wb_out.create_sheet(sheet_name)
    for col_idx, label in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label.replace("*", ""))
        if label.startswith("*"):
            cell.fill = yellow
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, label in enumerate(cols, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(label))
wb_out.save(output_path)

# ---------------------------------------------------------------------------
# 11. CSV output for Account and Location.
# ---------------------------------------------------------------------------

def format_csv_value(value):
    if value is None:
        return ''
    if isinstance(value, datetime.datetime):
        if (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0):
            return value.strftime('%Y-%m-%d')
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, datetime.date):
        return value.strftime('%Y-%m-%d')
    return value


for name, cols, rows in [("Account", account_columns, account_rows), ("Location", location_columns, location_rows)]:
    csv_path = output_dir / f"{name}.csv"
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([c.replace("*", "") for c in cols])
        for row in rows:
            writer.writerow([format_csv_value(row.get(c)) for c in cols])
