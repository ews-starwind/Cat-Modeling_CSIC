#!/usr/bin/env python3
"""
build_wind_tabs.py
==================

Reads the `CAPWindModeling` SOV export from a workbook and generates the
`Account` and `Location` tabs used for wind modeling.

Every output column corresponds to a YELLOW column in the `ModelingSOV`
working sheet of the original file. Those yellow cells are where the
mapping/derivation happens; the transformations they encode are reproduced
here in Python. Verified to reproduce the original tabs cell-for-cell.

Usage
-----
    python build_wind_tabs.py INPUT.xlsx [-o OUTPUT.xlsx] [--sheet CAPWindModeling]

If -o is omitted, the Account/Location tabs are written back into INPUT.xlsx
(any existing tabs with those names are replaced).

By default it ALSO writes Account.csv and Location.csv next to the output
.xlsx. Use --no-csv to skip them, or --csv-dir FOLDER to put them elsewhere.

Requires: openpyxl  ->  pip install openpyxl
"""

import csv
import sys
from collections import OrderedDict
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font


# ---------------------------------------------------------------------------
# Configuration — the modeling assumptions live here.
# ---------------------------------------------------------------------------

# Wind deductible scenarios expanded for EVERY account.
# The source WINDPCTDED is ignored on purpose: the model always produces all
# three scenarios. (tag, fraction) — the tag is appended to the account number
# and the fraction is written to WSSITEDED.
SCENARIOS = [("2pct", 0.02), ("3pct", 0.03), ("5pct", 0.05)]

# Constant "scheme" columns required by the modeling platform.
CNTRYSCHEME = "ISO2A"
CNTRYCODE   = "US"
BLDGSCHEME  = "FIRE"
OCCSCHEME   = "ATC"

# Policy definitions on the Account tab.
#   SCS = all-other-perils policy (POLICYTYPE 3, deductible -> BLANDEDAMT),
#         written once per account on the base scenario only.
#   WS  = wind/storm policy       (POLICYTYPE 2, deductible -> MINDEDAMT),
#         written once per account per scenario.
SCS_POLICYNUM, SCS_POLICYTYPE = "SCS", 3
WS_POLICYNUM,  WS_POLICYTYPE  = "WS", 2

INPUT_SHEET = "CAPWindModeling"

# Output headers (the yellow header cells in the original tabs).
ACCOUNT_HEADERS = [
    "ACCNTNUM", "ACCNTNAME", "INCEPTDATE", "EXPIREDATE",
    "POLICYNUM", "POLICYTYPE", "BLANDEDAMT", "MINDEDAMT",
]
LOCATION_HEADERS = [
    "ACCNTNUM", "ACCNTNAME", "LOCNUM", "STREETNAME", "CITY", "STATECODE",
    "POSTALCODE", "COUNTY", "CNTRYSCHEME", "CNTRYCODE", "BLDGSCHEME",
    "BLDGCLASS", "OCCSCHEME", "OCCTYPE", "NUMSTORIES", "YEARBUILT",
    "WSCV1VAL", "WSCV2VAL", "WSSITEDED", "TOCV1VAL", "TOCV2VAL", "FLOORAREA",
    "INCEPTDATE", "EXPIREDATE", "CONSTQUALI", "ROOFSYS", "ROOFAGE", "ROOFGEOM",
    "ROOFANCH", "CLADSYS", "FOUNDSYS", "ROOFEQUIP", "CLADRATE", "RESISTOPEN",
]

YELLOW = PatternFill("solid", fgColor="FFFF00")
HEADER_FONT = Font(name="Arial", bold=True)
BODY_FONT = Font(name="Arial")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _blank(v):
    """Treat empty strings as blank so they round-trip as truly empty cells."""
    return None if v == "" else v


def clean_name(s):
    """Strip commas from a text field (used for ACCNTNAME and STREETNAME).
    Commas break CSV columns and aren't wanted in the modeling output.
    e.g. 'PUERTO DEL SOL CONDOMINIUMS, INC.' -> 'PUERTO DEL SOL CONDOMINIUMS INC.'"""
    return str(s).replace(",", "") if s is not None else s


def read_source(ws):
    """Return (header->col map, list of rows grouped by account in file order)."""
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    accounts = OrderedDict()  # QUOTEID -> [row indices]
    for r in range(2, ws.max_row + 1):
        q = _blank(ws.cell(r, hdr["QUOTEID"]).value)
        if q in (None, ""):
            continue
        accounts.setdefault(q, []).append(r)
    return hdr, accounts


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------

def build_location(ws, hdr, accounts):
    def g(r, name):
        return _blank(ws.cell(r, hdr[name]).value)

    base_tag = SCENARIOS[0][0]
    rows = []
    for tag, frac in SCENARIOS:
        is_base = tag == base_tag
        for quoteid, src_rows in accounts.items():
            for r in src_rows:
                rows.append([
                    f"{quoteid}_{tag}",                 # ACCNTNUM
                    clean_name(g(r, "Named Insured")),  # ACCNTNAME
                    g(r, "BLDGNO"),                     # LOCNUM  (one loc per building)
                    clean_name(g(r, "STNAME")),         # STREETNAME (commas stripped)
                    g(r, "CITY"),                       # CITY
                    g(r, "STATE"),                      # STATECODE
                    g(r, "ZIP5"),                       # POSTALCODE
                    g(r, "COUNTY"),                     # COUNTY
                    CNTRYSCHEME, CNTRYCODE, BLDGSCHEME, # constant scheme cols
                    g(r, "CONSTCL"),                    # BLDGCLASS
                    OCCSCHEME,                          # OCCSCHEME
                    g(r, "OCCPCL"),                     # OCCTYPE
                    g(r, "NOSTORIES"),                  # NUMSTORIES
                    f"1/1/{g(r, 'YEARBUILT')}",         # YEARBUILT = CONCATENATE("1","/1/",YEARBUILT)
                    g(r, "LOCBLDREPL"),                 # WSCV1VAL  (building replacement cost)
                    g(r, "LOCCNTREPL"),                 # WSCV2VAL  (contents replacement cost)
                    frac,                               # WSSITEDED (scenario %)
                    # TOCV totals are written on the base scenario only:
                    g(r, "LOCBLDREPL") if is_base else None,   # TOCV1VAL
                    g(r, "LOCCNTREPL") if is_base else None,   # TOCV2VAL
                    g(r, "SQFEET"),                     # FLOORAREA
                    g(r, "Effective Date"),             # INCEPTDATE
                    g(r, "Expiration Date"),            # EXPIREDATE
                    g(r, "CONSTQUA"),                   # CONSTQUALI
                    g(r, "ROOFSYS"),                    # ROOFSYS
                    g(r, "ROOFAGE"),                    # ROOFAGE
                    g(r, "ROOFGEO"),                    # ROOFGEOM
                    g(r, "ROOFANC"),                    # ROOFANCH
                    g(r, "CLADSYS"),                    # CLADSYS
                    g(r, "FOUNDSYS"),                   # FOUNDSYS
                    g(r, "ROOFEQUI"),                   # ROOFEQUIP
                    g(r, "CLADRATE"),                   # CLADRATE
                    g(r, "RESISTOPEN"),                 # RESISTOPEN
                ])
    _write_sheet(ws.parent, "Location", LOCATION_HEADERS, rows)
    return LOCATION_HEADERS, rows


def build_account(ws, hdr, accounts):
    def g(r, name):
        return _blank(ws.cell(r, hdr[name]).value)

    base_tag = SCENARIOS[0][0]

    # One metadata record per account (taken from its first source row).
    meta = OrderedDict()
    for quoteid, src_rows in accounts.items():
        r0 = src_rows[0]
        meta[quoteid] = dict(
            name=clean_name(g(r0, "Named Insured")),
            eff=g(r0, "Effective Date"),
            exp=g(r0, "Expiration Date"),
            dedu=g(r0, "LOCBLDDEDU"),   # flat minimum deductible for this account
        )

    rows = []
    # SCS all-other-perils policy: base scenario only, deductible -> BLANDEDAMT
    for quoteid, m in meta.items():
        rows.append([f"{quoteid}_{base_tag}", m["name"], m["eff"], m["exp"],
                     SCS_POLICYNUM, SCS_POLICYTYPE, m["dedu"], None])
    # WS wind policy: every scenario, deductible -> MINDEDAMT
    for tag, _frac in SCENARIOS:
        for quoteid, m in meta.items():
            rows.append([f"{quoteid}_{tag}", m["name"], m["eff"], m["exp"],
                         WS_POLICYNUM, WS_POLICYTYPE, None, m["dedu"]])

    _write_sheet(ws.parent, "Account", ACCOUNT_HEADERS, rows)
    return ACCOUNT_HEADERS, rows


def _write_sheet(wb, name, headers, rows):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = YELLOW
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(i, c, v)
            cell.font = BODY_FONT
    return ws


def _write_csv(path, headers, rows):
    # newline="" is the documented way to avoid blank lines on Windows.
    # Blank cells (None) are written as empty fields.
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for row in rows:
            w.writerow(["" if v is None else v for v in row])


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    # Anchor all paths to the directory that physically contains this program,
    # so it keeps working when the executable is copied to a new location.
    #   - Compiled (Nuitka sets __compiled__, PyInstaller sets sys.frozen):
    #     use the directory of the actual .exe (sys.executable). NOT the cwd,
    #     which is only correct when the program is launched from its own folder.
    #   - Plain script: use the directory of this source file.
    if getattr(sys, "frozen", False) or "__compiled__" in globals():
        base_dir = Path(sys.executable).resolve().parent
    else:
        base_dir = Path(__file__).resolve().parent

    data_files = list((base_dir / "data").glob("*.xlsx"))
    if len(data_files) != 1:
        raise SystemExit(f"Expected exactly one .xlsx file in data/, found: {data_files}")
    input_path = data_files[0]
    out = base_dir / "output" / "WindModeling_Output.xlsx"
    csv_dir = out.parent

    # data_only=True reads the cached VALUES of any source formulas.
    wb = openpyxl.load_workbook(input_path, data_only=True)

    src = wb[wb.sheetnames[0]]
    hdr, accounts = read_source(src)
    if not accounts:
        raise SystemExit("No data rows found (empty QUOTEID column).")

    loc_headers, loc_rows = build_location(src, hdr, accounts)
    acc_headers, acc_rows = build_account(src, hdr, accounts)

    csv_dir.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    acc_csv = csv_dir / "Account.csv"
    loc_csv = csv_dir / "Location.csv"
    _write_csv(acc_csv, acc_headers, acc_rows)
    _write_csv(loc_csv, loc_headers, loc_rows)
    csv_paths = [acc_csv, loc_csv]

    n_bldgs = sum(len(v) for v in accounts.values())
    print(f"Done -> {out}")
    print(f"  Accounts:  {len(accounts)}")
    print(f"  Buildings: {n_bldgs}")
    print(f"  Scenarios: {', '.join(t for t, _ in SCENARIOS)}")
    print(f"  Location rows: {n_bldgs * len(SCENARIOS)}")
    print(f"  Account rows:  {len(accounts) * (1 + len(SCENARIOS))}")
    for p in csv_paths:
        print(f"  CSV -> {p}")


if __name__ == "__main__":
    main()
