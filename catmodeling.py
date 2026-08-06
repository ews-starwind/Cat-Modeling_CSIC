import python_calamine
import pandas as pd
import numpy as np
from pathlib import Path
from typing import cast
import shutil
from openpyxl.styles import PatternFill

base_path = Path.cwd()
excel_files = list((base_path / "data").glob("*.xlsx"))
if len(excel_files) != 1:
    raise SystemExit(f"Expected exactly one .xlsx in data/, found: {excel_files}")
raw_data = excel_files[0]

raw_data_df = pd.read_excel(raw_data, engine="calamine")
# raw_data_df.head(10)
# Drop any unnamed columns (blank headers); no-op when there are none
unnamed_cols = [c for c in raw_data_df.columns if str(c).startswith("Unnamed")]
if unnamed_cols:
    raw_data_df = raw_data_df.drop(columns=unnamed_cols)

RENAME_COLS = {
    'MCONAME' : 'MCONAME',
    'Agent Name' : 'Agent Name',
    'Named Insured' : 'Named Insured',  #get ACCNTNAME from name insured with commas stripped for proper CSV parsing
    'QUOTEID' : 'QUOTEID',  # ACCNTNUM gets value from this column
    'LOCNO' : 'LOCNO',
    'BLDGNO' : 'BLDGNO',    #get LOCNUM column from this column
    'STNAME' : 'STNAME',    #get STREETNAME column from this column
    'CITY' : 'CITY',        #duplicate CITY column? --this probably matters when we get only the yellow columns for the CSV output so that the CSV parses correctly when inputted into the model
    'STATE' : 'STATE',      #get STATECODE column from this column
    'ZIP5' : 'ZIP5',        #get POSTALCODE column from this column
    'COUNTY' : 'COUNTY',    #duplicate COUNTY column; other columns: CNTRYSCHEME == ISO2A, CNTRYCODE == US
    'CONSTCL' : 'CONSTCL',  #after this we have BLDGSCHEME == FIRE and BLDGCLASS which comes from this column
    'OCCPCL' : 'OCCPCL',    #after this we have OCCSCHEME == ATC and OCCTYPE which comes from this column
    'NOSTORIES' : 'NOSTORIES',  #get NUMSTORIES column from this column
    'YEARBUILT' : 'YEARBUILT',  #duplicate YEARBUILT column
    'RMSLOB' : 'RMSLOB',
    'WINDPCTDED' : 'WINDPCTDED',
    'LOCBLKLIMIT' : 'LOCBLKLIMIT',
    'LOCBLKDEDU' : 'LOCBLKDEDU',
    'LOCBLDREPL' : 'LOCBLDREPL',    #get CV1VAL column from this column
    'LOCBLDLIMT' : 'LOCBLDLIMT',
    'LOCBLDDEDU' : '*LOCBLDDEDU',
    'LOCCNTREPL' : 'LOCCNTREPL',    #get CV2VAL column from this column
    'COCCNTLIMT' : 'COCCNTLIMT',
    'LOCCNTDEDU' : 'LOCCNTDEDU',
    'LOCBIRC' : 'LOCBIRC',
    'LOCBILIMIT' : 'LOCBILIMIT',
    'LOCBIDEDU' : 'LOCBIDEDU',
    'LOCBLDPREMAOP' : 'LOCBLDPREMAOP',
    'LOCBLDPREMWIND' : 'LOCBLDPREMWIND',
    'LOCCNTPREMAOP' : 'LOCCNTPREMAOP',
    'LOCCNTPREMWIND' : 'LOCCNTPREMWIND',
    'LOCBIPREM' : 'LOCBIPREM',
    'SQFEET' : 'SQFEET',    #get FLOORAREA column from this column
    'RATINGTERR' : 'RATINGTERR',
    'SPRINKLER' : 'SPRINKLER',
    'PROTCLASS' : 'PROTCLASS',
    'DESCRIPTION' : 'DESCRIPTION',
    'Effective Date' : 'Effective Date',    #get INCEPTDATE column from this column
    'Expiration Date' : 'Expiration Date',  #get EXPIREDATE column from this column
    'CONSTQUA' : '*CONSTQUALI',
    'ROOFSYS' : '*ROOFSYS',
    'ROOFAGE' : '*ROOFAGE',
    'ROOFGEO' : '*ROOFGEOM',
    'ROOFANC' : '*ROOFANCH',
    'CLADSYS' : '*CLADSYS',
    'FOUNDSYS' : '*FOUNDSYS',
    'ROOFEQUI' : '*ROOFEQUIP',
    'CLADRATE' : '*CLADRATE',
    'RESISTOPEN' : '*RESISTOPEN'
}

SCENARIOS = [
    ('2pct', 'SCS', 3, 0.02),
    ('2pct', 'WS', 2, 0.02),
    ('3pct', 'WS', 2, 0.03),
    ('5pct', 'WS', 2, 0.05)
]

modeling = raw_data_df.rename(columns = RENAME_COLS)

named_insured_loc = cast(int, modeling.columns.get_loc('Named Insured'))
modeling.insert(named_insured_loc + 1, '*ACCNTNAME', modeling['Named Insured'].str.replace(',', '', regex=False))
quoteid_loc = cast(int, modeling.columns.get_loc('QUOTEID'))
modeling.insert(quoteid_loc + 1, '*ACCNTNUM', modeling['QUOTEID'])
bldgno_loc = cast(int, modeling.columns.get_loc('BLDGNO'))
modeling.insert(bldgno_loc + 1, '*LOCNUM', modeling['BLDGNO'])
county_loc = cast(int, modeling.columns.get_loc('COUNTY'))
modeling.insert(county_loc + 1, '*STREETNAME', modeling['STNAME'].str.replace(',', '', regex=False))
modeling.insert(county_loc + 2, '*CITY', modeling['CITY'].str.replace(',', '', regex=False))
modeling.insert(county_loc + 3, '*STATECODE', modeling['STATE'])
modeling.insert(county_loc + 4, '*POSTALCODE', modeling['ZIP5'])
modeling.insert(county_loc + 5, '*COUNTY', modeling['COUNTY'].str.replace(',', '', regex=False))
modeling.insert(county_loc + 6, '*CNTRYSCHEME', 'ISO2A')
modeling.insert(county_loc + 7, '*CNTRYCODE', 'US')
constcl_loc = cast(int, modeling.columns.get_loc('CONSTCL'))
modeling.insert(constcl_loc + 1, '*BLDGSCHEME', 'FIRE')
modeling.insert(constcl_loc + 2, '*BLDGCLASS', modeling['CONSTCL'])
occpcl_loc = cast(int, modeling.columns.get_loc('OCCPCL'))
modeling.insert(occpcl_loc + 1, '*OCCSCHEME', 'ATC')
modeling.insert(occpcl_loc + 2, '*OCCTYPE', modeling['OCCPCL'])
nostories_loc = cast(int, modeling.columns.get_loc('NOSTORIES'))
modeling.insert(nostories_loc + 1, '*NUMSTORIES', modeling['NOSTORIES'])
yearbuilt_loc = cast(int, modeling.columns.get_loc('YEARBUILT'))
modeling.insert(yearbuilt_loc + 1, '*YEARBUILT', '1/1/' + modeling['YEARBUILT'].astype('Int64').astype(str).str.replace(',', '', regex=False))
locbldrepl_loc = cast(int, modeling.columns.get_loc('LOCBLDREPL'))
modeling.insert(locbldrepl_loc, '*CV1VAL', modeling['LOCBLDREPL'])
loccntrepl_loc = cast(int, modeling.columns.get_loc('LOCCNTREPL'))
modeling.insert(loccntrepl_loc, '*CV2VAL', modeling['LOCCNTREPL'])
sqfeet_loc = cast(int, modeling.columns.get_loc('SQFEET'))
modeling.insert(sqfeet_loc + 1, '*FLOORAREA', modeling['SQFEET'])
expiration_loc = cast(int, modeling.columns.get_loc('Expiration Date'))
modeling.insert(expiration_loc + 1, '*INCEPTDATE', modeling['Effective Date'])
modeling.insert(expiration_loc + 2, '*EXPIREDATE', modeling['Expiration Date'])


acct_temp = (modeling[['*ACCNTNUM','*ACCNTNAME','*INCEPTDATE','*EXPIREDATE','*LOCBLDDEDU']]
              .drop_duplicates(subset = ['*ACCNTNUM']))

location_temp = modeling[[col for col in modeling.columns if col.startswith('*')]]

acct_scenario_dfs = []
loc_scenario_dfs = []

for scenario in SCENARIOS:
    pct, pnum, ptype, prob = scenario

    acct_scenario_df = acct_temp.copy()
    acct_scenario_df['*ACCNTNUM'] = acct_scenario_df['*ACCNTNUM'].astype(str) + '_' + pct
    acct_scenario_df.insert(acct_scenario_df.columns.get_loc('*EXPIREDATE') + 1, '*POLICYNUM', pnum)
    acct_scenario_df.insert(acct_scenario_df.columns.get_loc('*EXPIREDATE') + 2, '*POLICYTYPE', ptype)

    loc_scenario_df = location_temp.copy()
    loc_scenario_df['*ACCNTNUM'] = loc_scenario_df['*ACCNTNUM'].astype(str) + '_' + pct
    loc_scenario_df.insert(loc_scenario_df.columns.get_loc('*CV1VAL') + 1, '*WSSITEDED', prob)

    if pnum == 'SCS':
        acct_scenario_df['*BLANDEDAMT'] = acct_scenario_df['*LOCBLDDEDU']
        acct_scenario_df['*MINDEDAMT'] = np.nan

        loc_scenario_df.insert(loc_scenario_df.columns.get_loc('*CV2VAL') + 1, '*TOCV1VAL', loc_scenario_df['*CV1VAL'])
        loc_scenario_df.insert(loc_scenario_df.columns.get_loc('*CV2VAL') + 2, '*TOCV2VAL', loc_scenario_df['*CV2VAL'])
        loc_scenario_df.insert(loc_scenario_df.columns.get_loc('*CV1VAL') + 1, '*WSCV1VAL', np.nan)
        loc_scenario_df.insert(loc_scenario_df.columns.get_loc('*CV1VAL') + 2, '*WSCV2VAL', np.nan)
        
    else:
        acct_scenario_df['*BLANDEDAMT'] = np.nan
        acct_scenario_df['*MINDEDAMT'] = acct_scenario_df['*LOCBLDDEDU']

        loc_scenario_df.insert(loc_scenario_df.columns.get_loc('*CV2VAL') + 1, '*TOCV1VAL', np.nan)
        loc_scenario_df.insert(loc_scenario_df.columns.get_loc('*CV2VAL') + 2, '*TOCV2VAL', np.nan)
        loc_scenario_df.insert(loc_scenario_df.columns.get_loc('*CV1VAL') + 1, '*WSCV1VAL', loc_scenario_df['*CV1VAL'])
        loc_scenario_df.insert(loc_scenario_df.columns.get_loc('*CV1VAL') + 2, '*WSCV2VAL', loc_scenario_df['*CV2VAL'])

    loc_scenario_df = (loc_scenario_df
                       .drop(columns=['*CV1VAL', '*CV2VAL']))

    acct_scenario_dfs.append(acct_scenario_df)
    loc_scenario_dfs.append(loc_scenario_df)

account_df = pd.concat(acct_scenario_dfs, ignore_index=True).drop(columns=['*LOCBLDDEDU'])

location_df = pd.concat(loc_scenario_dfs, ignore_index=True).drop(columns=['*LOCBLDDEDU'])
location_cols = location_df.columns.tolist()
name_idx, num_idx = location_cols.index('*ACCNTNAME'), location_cols.index('*ACCNTNUM')
location_cols[name_idx], location_cols[num_idx] = location_cols[num_idx], location_cols[name_idx]
location_df = (location_df
               .groupby(['*ACCNTNUM', '*LOCNUM'], as_index=False, sort=False)
               .first())[location_cols]

output_dir = base_path / "output"
output_dir.mkdir(exist_ok=True)
output_path = output_dir / "WindModeling_Output.xlsx"

# CAPWindModeling is the raw import shown as-is. Copy the source workbook verbatim so
# its original formatting (header styling, fonts, borders, column widths) is preserved
# exactly, then append the derived sheets to it.
shutil.copyfile(raw_data, output_path)

derived_sheets = [
    ("ModelingSOV", modeling),
    ("Account",     account_df),
    ("Location",    location_df),
]

with pd.ExcelWriter(output_path, engine="openpyxl", mode="a") as writer:
    yellow = PatternFill(fill_type="solid", fgColor="FFFF00")
    for sheet_name, df in derived_sheets:
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=1)
        ws = writer.sheets[sheet_name]
        for col, label in enumerate(df.columns, start=1):
            label = str(label)
            cell = ws.cell(row=1, column=col, value=label.replace("*", ""))
            if label.startswith("*"):
                cell.fill = yellow

for name, df in [("Account", account_df), ("Location", location_df)]:
    out = df.copy()
    out.columns = [str(c).replace("*", "") for c in out.columns]
    str_cols = out.select_dtypes(include='object').columns
    out[str_cols] = out[str_cols].apply(lambda s: s.astype(str).str.replace(',', ''))
    out.to_csv(output_dir / f"{name}.csv", index=False, encoding="utf-8-sig")